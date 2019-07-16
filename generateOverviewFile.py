#! python3
# generateOverviewFile.py

# Generates a single excel workbook, the first sheet being an overview
# and each subsequent sheet containing all data for a single employee

# import excel and os
import openpyxl, os

# placeholder variables
openRow = 0

# open overview book
overview = openpyxl.load_workbook('overview.xlsx')

# create employees list from the sheet names
employees = overview.sheetnames

print('Please wait while the overview sheet is updated...')

# loop through all excel files that do not contain "overview"
# os.chdir([path to your employee records goes here])
for filename in os.listdir():
    if filename.endswith('.xlsx') and not 'overview' in filename:

        # open the excel file and select the sheet
        file = openpyxl.load_workbook(filename)
        sheet = file.active

        # check to see if employee is in employees
        emp = sheet['B1'].value
        if emp not in employees:

            # add emp to employees
            employees.append(emp)

            # genereate a new worksheet for emp in overview book
            empSheet = overview.create_sheet(emp)

            # add header to new sheet
            empSheet['A1'] = 'Week'
            empSheet['B1'] = 'Hours'

        else:
            empSheet = overview[emp]

        # generate week and hours plus linked cell IDs
        week = sheet['D1'].value
        hours = sheet['A2'].value
        row = week + 1
        weekCellID = 'A%s' % row
        hoursCellID = 'B%s' % row

        # print week and hours in relevant row
        empSheet[weekCellID].value = week
        empSheet[hoursCellID].value = hours             

        # close temp file
        file.close()

# select overview sheet
sheet = overview['Overview']

# format sheet
sheet['A1'].value = 'Type week number below:'
row = 1
column = 2
for emp in employees:
    if emp != 'Overview':
        head = sheet.cell(row = row, column = column)
        head.value = emp
        formula = """=VLOOKUP($A$2, INDIRECT("'"&%s&"'!A:B"), 2, FALSE)""" % head.coordinate
        sheet.cell(row = row + 1, column = column).value = formula
        column += 1
 

# save and close overview
overview.save('overview.xlsx')
overview.close

print('Overview generation complete!')
        
            
        
